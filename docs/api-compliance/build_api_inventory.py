import json
import os
import re
import ssl
import urllib.error
import urllib.request
import xml.etree.ElementTree as ET
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo


BASE_URL = os.environ.get("CCODER_API_BASE_URL", "https://localhost:7099")
OUTPUT_DIR = Path(__file__).resolve().parent
RAW_DIR = OUTPUT_DIR / "raw"
API_ROOT_PATH = "/Api"
CORE_SWAGGER_PATH = "/swagger/Core/swagger.json"
CORE_PACKAGE_VERSION = os.environ.get("CCODER_CORE_PACKAGE_VERSION", "Not recorded")
CAPTURE_PROFILE = os.environ.get("CCODER_API_CAPTURE_PROFILE", "Configured aggregate")

HTTP_METHODS = {"get", "head", "post", "put", "patch", "delete", "options", "trace"}


def fetch(path):
    context = ssl._create_unverified_context()
    request = urllib.request.Request(BASE_URL + path, headers={"Accept": "*/*"})
    with urllib.request.urlopen(request, context=context, timeout=30) as response:
        return response.read(), response.headers.get_content_type(), response.status


def success_expectation(method, path, operation):
    method = method.upper()
    text = f"{path} {operation.get('operationId', '')} {' '.join(operation.get('tags', []))}".lower()
    if method in {"GET", "HEAD"}:
        return "200", "RFC 9110 §§9.3.1-9.3.2; provisional cCoder policy"
    if method == "DELETE":
        return "204", "RFC 9110 §15.3.5; provisional cCoder policy"
    if method == "POST" and any(word in text for word in ("create", "register", "invite")):
        return "201", "RFC 9110 §15.3.2; confirm resource-creation semantics"
    if method == "POST":
        return "200", "RFC 9110 §9.3.3; operation-specific, requires review"
    if method in {"PUT", "PATCH"}:
        return "200 or 204", "Representation-dependent; choose cCoder policy"
    if method == "OPTIONS":
        return "200 or 204", "RFC 9110 §9.3.7"
    return "Review", "Operation-specific"


def add_table(ws, name):
    if ws.max_row < 2 or ws.max_column < 1:
        return
    table = Table(displayName=name, ref=f"A1:{ws.cell(ws.max_row, ws.max_column).coordinate}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False,
        showRowStripes=True, showColumnStripes=False)
    ws.add_table(table)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = table.ref


def format_sheet(ws, widths):
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="top", wrap_text=True)
    for column, width in widths.items():
        ws.column_dimensions[column].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def main():
    RAW_DIR.mkdir(parents=True, exist_ok=True)

    for pattern in ("api-root.json", "swagger-*.json", "odata-*.xml"):
        for stale_file in RAW_DIR.glob(pattern):
            stale_file.unlink()

    generated = datetime.now(timezone.utc).isoformat()

    confirmed_by_operation = {}
    advertised_operation_count = 0
    advertised_codes = set()
    sources = []
    context_summary = {}
    api_root_bytes, api_root_type, api_root_status = fetch(API_ROOT_PATH)
    (RAW_DIR / "api-root.json").write_bytes(api_root_bytes)
    api_root = json.loads(api_root_bytes)
    sources.append(["API root", API_ROOT_PATH, api_root_status, api_root_type, len(api_root_bytes), generated, "Captured"])
    contexts = [item for item in api_root.get("value", []) if item.get("kind") == "Context"]
    swagger_documents = [("Core", CORE_SWAGGER_PATH)] + [
        (item["name"], item["swaggerDef"]) for item in contexts if item.get("swaggerDef")
    ]
    for definition, swagger_path in swagger_documents:
        swagger_bytes, swagger_type, swagger_status = fetch(swagger_path)
        safe_name = definition.replace("/", "-")
        (RAW_DIR / f"swagger-{safe_name}.json").write_bytes(swagger_bytes)
        sources.append(["Swagger", swagger_path, swagger_status, swagger_type, len(swagger_bytes), generated, "Captured"])
        swagger = json.loads(swagger_bytes)
        operation_count = 0
        for path, path_item in swagger.get("paths", {}).items():
            for method, operation in path_item.items():
                if method.lower() not in HTTP_METHODS:
                    continue
                operation_count += 1
                advertised_operation_count += 1
                identity = (path.lower(), method.lower())
                responses = operation.get("responses", {})
                advertised = ", ".join(responses.keys())
                advertised_codes.update(responses.keys())
                expected, basis = success_expectation(method, path, operation)
                if identity in confirmed_by_operation:
                    existing = confirmed_by_operation[identity]
                    existing[3] = ", ".join(sorted(set(existing[3].split(", ") + [definition])))
                    existing[6] = "\n".join(sorted(set(existing[6].split("\n") + [swagger_path])))
                    continue
                confirmed_by_operation[identity] = [
                    "Web", path, method.upper(), definition,
                    operation.get("operationId", ""), "Swagger confirmed", swagger_path,
                    advertised, expected, basis, "", "", "", "", "Not assessed"
                ]
        context_summary[definition] = {
            "swagger_path": swagger_path,
            "swagger_status": swagger_status,
            "operations": operation_count,
            "metadata_path": "Not applicable" if definition == "Core" else f"/Api/{definition}/$metadata",
            "metadata_status": "Not applicable" if definition == "Core" else "Not requested",
            "entity_sets": 0,
        }

    confirmed = list(confirmed_by_operation.values())

    model_rows = []
    candidate_rows = []
    edm_ns = "http://docs.oasis-open.org/odata/ns/edm"
    metadata_services = {item["name"] for item in contexts}
    for row in confirmed:
        match = re.fullmatch(r"/Api/([^/]+)/\$metadata", row[1], flags=re.IGNORECASE)
        if match:
            metadata_services.add(match.group(1))
    for service in sorted(metadata_services):
        path = f"/Api/{service}/$metadata"
        try:
            xml_bytes, content_type, status = fetch(path)
        except urllib.error.HTTPError as error:
            sources.append(["OData metadata", path, error.code, "", 0, generated, "Unavailable"])
            context_summary[service]["metadata_status"] = error.code
            continue
        (RAW_DIR / f"odata-{service}.xml").write_bytes(xml_bytes)
        sources.append(["OData metadata", path, status, content_type, len(xml_bytes), generated, "Captured"])
        context_summary[service]["metadata_status"] = status
        root = ET.fromstring(xml_bytes)
        schemas = root.findall(f".//{{{edm_ns}}}Schema")
        type_keys = {}
        for schema in schemas:
            namespace = schema.get("Namespace", "")
            for entity_type in schema.findall(f"{{{edm_ns}}}EntityType"):
                keys = [ref.get("Name", "") for ref in entity_type.findall(f"{{{edm_ns}}}Key/{{{edm_ns}}}PropertyRef")]
                type_keys[f"{namespace}.{entity_type.get('Name', '')}"] = ", ".join(keys)
        for container in root.findall(f".//{{{edm_ns}}}EntityContainer"):
            for entity_set in container.findall(f"{{{edm_ns}}}EntitySet"):
                name = entity_set.get("Name", "")
                entity_type = entity_set.get("EntityType", "")
                key = type_keys.get(entity_type, "")
                model_rows.append([service, "EntitySet", name, entity_type, key, path])
                context_summary[service]["entity_sets"] += 1
                base = f"/Api/{service}/{name}"
                patterns = [
                    (base, "GET", "200", "Read collection"),
                    (base, "POST", "201", "Create entity"),
                    (base + "({key})", "GET", "200", "Read entity"),
                    (base + "({key})", "PATCH", "200 or 204", "Update entity"),
                    (base + "({key})", "PUT", "200 or 204", "Replace/upsert entity"),
                    (base + "({key})", "DELETE", "204", "Delete entity"),
                ]
                for candidate_path, method, expected, scenario in patterns:
                    candidate_rows.append([
                        "Web", candidate_path, method, service, "", "OData inferred candidate", path,
                        "Not advertised", expected, "OData convention; implementation must be verified",
                        scenario, "", "", "", "Discovery required"
                    ])
            for singleton in container.findall(f"{{{edm_ns}}}Singleton"):
                name = singleton.get("Name", "")
                entity_type = singleton.get("Type", "")
                model_rows.append([service, "Singleton", name, entity_type, type_keys.get(entity_type, ""), path])

    core_metadata_path = "/Api/Core/$metadata"
    try:
        core_metadata_bytes, core_metadata_type, core_metadata_status = fetch(core_metadata_path)
        sources.append([
            "Core OData negative check", core_metadata_path, core_metadata_status,
            core_metadata_type, len(core_metadata_bytes), generated, "Unexpectedly available"
        ])
    except urllib.error.HTTPError as error:
        core_metadata_status = error.code
        sources.append([
            "Core OData negative check", core_metadata_path, error.code,
            "", 0, generated, "Correctly unavailable" if error.code == 404 else "Unavailable"
        ])

    headers = [
        "Host", "Relative Path", "HTTP Verb", "Domain / Tag", "Operation ID",
        "Discovery Confidence", "Source Document", "Advertised Response Codes",
        "Provisional Good Response Code", "Standards / Policy Basis", "Scenario",
        "Required Response Headers", "Response Body Requirement", "Required Error Codes",
        "Test Coverage Status"
    ]
    wb = Workbook()
    ws = wb.active
    ws.title = "Confirmed Endpoints"
    ws.append(headers)
    for row in sorted(confirmed, key=lambda r: (r[1], r[2])):
        ws.append(row)
    add_table(ws, "ConfirmedEndpoints")
    format_sheet(ws, {"A": 12, "B": 48, "C": 12, "D": 24, "E": 34, "F": 22, "G": 34,
                      "H": 24, "I": 26, "J": 48, "K": 24, "L": 30, "M": 30, "N": 30, "O": 20})

    ws = wb.create_sheet("OData Candidates")
    ws.append(headers)
    for row in sorted(candidate_rows, key=lambda r: (r[3], r[1], r[2])):
        ws.append(row)
    add_table(ws, "ODataCandidates")
    format_sheet(ws, {"A": 12, "B": 52, "C": 12, "D": 24, "E": 20, "F": 24, "G": 38,
                      "H": 20, "I": 26, "J": 48, "K": 24, "L": 30, "M": 30, "N": 30, "O": 20})

    ws = wb.create_sheet("OData Model")
    ws.append(["Service", "Kind", "Name", "Entity Type", "Declared Key", "Metadata Source"])
    for row in sorted(model_rows):
        ws.append(row)
    add_table(ws, "ODataModel")
    format_sheet(ws, {"A": 24, "B": 16, "C": 30, "D": 55, "E": 24, "F": 42})

    ws = wb.create_sheet("Source Documents")
    ws.append(["Kind", "Relative URL", "HTTP Status", "Content Type", "Bytes", "Captured UTC", "Result"])
    for row in sources:
        ws.append(row)
    add_table(ws, "SourceDocuments")
    format_sheet(ws, {"A": 20, "B": 45, "C": 14, "D": 30, "E": 14, "F": 30, "G": 18})

    duplicate_operations = [
        row for row in confirmed
        if len([name for name in row[3].split(", ") if name]) > 1
    ]

    response_coverage = {
        code: sum(
            1 for row in confirmed
            if code in [item.strip() for item in (row[7] or "").split(",")]
        )
        for code in ("200", "201", "204", "400", "401", "403", "404", "409", "412", "415", "500")
    }
    only_200_operations = sum(1 for row in confirmed if (row[7] or "").strip() == "200")

    ws = wb.create_sheet("Contract Summary")
    ws.append(["Measure", "Value", "Assessment"])
    contract_rows = [
        ["Core package version", CORE_PACKAGE_VERSION, "Capture provenance"],
        ["Capture profile", CAPTURE_PROFILE, "Capture provenance"],
        ["Swagger documents", len(swagger_documents), "Core plus configured child contexts"],
        ["Advertised operations", advertised_operation_count, "Before path and verb de-duplication"],
        ["Unique path and verb pairs", len(confirmed), "Aggregate contract surface"],
        ["Duplicate path and verb pairs", len(duplicate_operations), "Expected: 0"],
        ["Operations advertising only 200", only_200_operations, "Expected: 0"],
        ["Core OData metadata status", core_metadata_status, "Expected: 404; Core owns no OData context"],
    ]
    contract_rows.extend([
        [f"Operations advertising {code}", count, "Response contract coverage"]
        for code, count in response_coverage.items()
    ])
    for row in contract_rows:
        ws.append(row)
    add_table(ws, "ContractSummary")
    format_sheet(ws, {"A": 42, "B": 34, "C": 58})

    ws = wb.create_sheet("API Contexts")
    ws.append([
        "Context", "Swagger URL", "Swagger Status", "Advertised Operations",
        "OData Metadata URL", "Metadata Status", "Entity Sets"
    ])
    for context_name, summary in sorted(context_summary.items()):
        ws.append([
            context_name,
            summary["swagger_path"],
            summary["swagger_status"],
            summary["operations"],
            summary["metadata_path"],
            summary["metadata_status"],
            summary["entity_sets"],
        ])
    add_table(ws, "ApiContexts")
    format_sheet(ws, {"A": 34, "B": 48, "C": 16, "D": 24, "E": 56, "F": 18, "G": 16})

    ws = wb.create_sheet("Read Me")
    notes = [
        ["Item", "Value"],
        ["Purpose", "Initial inventory for planning RFC 9110 and OData compliance tests."],
        ["Core package version", CORE_PACKAGE_VERSION],
        ["Capture profile", CAPTURE_PROFILE],
        ["Live host", BASE_URL],
        ["Generated UTC", generated],
        ["Confirmed endpoints", len(confirmed)],
        ["OData model members", len(model_rows)],
        ["OData candidate operations", len(candidate_rows)],
        ["Swagger operations advertised by multiple definitions", len(duplicate_operations)],
        ["Important", "OData candidates are inferred from entity sets. They are not proof that the route/verb is implemented."],
        ["Good response codes", "Provisional planning values only. They require an agreed cCoder API policy and scenario-specific review."],
        ["Next step", "Verify candidates against routing, then define good/error scenarios, headers, body rules, and existing test coverage."],
        ["RFC source", "https://www.rfc-editor.org/rfc/rfc9110.html"],
        ["OData source", "https://docs.oasis-open.org/odata/odata/v4.01/odata-v4.01-part1-protocol.html"],
    ]
    for row in notes:
        ws.append(row)
    format_sheet(ws, {"A": 28, "B": 110})
    ws.freeze_panes = "A2"

    output = Path(os.environ.get(
        "CCODER_API_INVENTORY_OUTPUT",
        str(OUTPUT_DIR / "cCoder.Core API Inventory.xlsx")))
    wb.save(output)
    checked = load_workbook(output, read_only=True, data_only=False)
    assert len(checked.sheetnames) == 7
    assert checked["Confirmed Endpoints"].max_row == len(confirmed) + 1
    print(json.dumps({
        "output": str(output), "confirmed_operations": len(confirmed),
        "odata_model_members": len(model_rows), "odata_candidates": len(candidate_rows),
        "advertised_response_codes": sorted(advertised_codes), "sources": len(sources),
        "swagger_documents": len(swagger_documents),
        "advertised_operations": advertised_operation_count,
        "duplicate_path_verb_pairs": len(duplicate_operations),
        "only_200_operations": only_200_operations,
        "core_metadata_status": core_metadata_status,
        "response_coverage": response_coverage,
    }, indent=2))


if __name__ == "__main__":
    main()
